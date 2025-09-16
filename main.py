from fastapi import FastAPI
from fastapi.responses import FileResponse
import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import logging

# ---------- 配置 ----------
logging.basicConfig(level=logging.INFO)

app = FastAPI()
DATA_DIR = "data"
os.makedirs(DATA_DIR, exist_ok=True)
MAX_CITIES = 10
COUNTRY_FILE = "countries.txt"

# 国家信息: 语言, 时区, 洲
COUNTRY_INFO = {
    "中国": ("汉语", "UTC+8", "亚洲"),
    "美国": ("英语", "UTC-5 至 UTC-10", "北美洲"),
    "埃及": ("阿拉伯语", "UTC+2", "非洲"),
    "尼日利亚": ("英语", "UTC+1", "非洲"),
    "法国": ("法语", "UTC+1", "欧洲"),
    "日本": ("日语", "UTC+9", "亚洲"),
    "澳大利亚": ("英语", "UTC+8 至 UTC+10", "大洋洲"),
    # 可自行补全其他国家
}

# ---------- 读取国家列表 ----------
def read_countries():
    countries = []
    if os.path.exists(COUNTRY_FILE):
        with open(COUNTRY_FILE, "r", encoding="utf-8") as f:
            for line in f:
                country = line.strip()
                if country:
                    countries.append(country)
    return countries

# ---------- 抓取城市和人口 ----------
def fetch_country_cities(country_name):
    """从多个互联网网站抓取前10城市及人口"""

    # ---------------- 中文 Wikipedia ----------------
    urls_zh = [
        f"https://zh.wikipedia.org/wiki/{country_name}城市列表",
        f"https://zh.wikipedia.org/wiki/{country_name}"
    ]
    for url in urls_zh:
        try:
            resp = requests.get(url, timeout=15)
            resp.encoding = "utf-8"
            soup = BeautifulSoup(resp.text, "html.parser")
            tables = soup.find_all("table", {"class": ["wikitable", "wikitable sortable"]})
            for table in tables:
                rows = table.find_all("tr")[1:]
                city_pop_list = []
                for row in rows:
                    cols = row.find_all(["td", "th"])
                    if len(cols) < 2:
                        continue
                    city = cols[0].get_text(strip=True)
                    pop_text = cols[1].get_text(strip=True).replace(",", "").split()[0]
                    try:
                        pop = int(pop_text)
                    except:
                        continue
                    city_pop_list.append((city, pop))
                    if len(city_pop_list) >= MAX_CITIES:
                        break
                if city_pop_list:
                    return city_pop_list
        except:
            continue

    # ---------------- 英文 Wikipedia ----------------
    url_en = f"https://en.wikipedia.org/wiki/List_of_cities_in_{country_name.replace(' ', '_')}"
    try:
        resp = requests.get(url_en, timeout=15)
        soup = BeautifulSoup(resp.text, "html.parser")
        tables = soup.find_all("table", {"class": ["wikitable", "wikitable sortable"]})
        for table in tables:
            headers = [th.get_text(strip=True).lower() for th in table.find_all("th")]
            if not any("population" in h for h in headers):
                continue
            rows = table.find_all("tr")[1:]
            city_pop_list = []
            for row in rows:
                cols = row.find_all("td")
                if len(cols) < 2:
                    continue
                city = cols[0].get_text(strip=True)
                pop_text = cols[1].get_text(strip=True).replace(",", "").split()[0]
                try:
                    pop = int(pop_text)
                except:
                    pop = 0
                city_pop_list.append((city, pop))
                if len(city_pop_list) >= MAX_CITIES:
                    break
            if city_pop_list:
                return city_pop_list
    except:
        pass

    # ---------------- World Population Review ----------------
    try:
        url_wp = f"https://worldpopulationreview.com/countries/{country_name.replace(' ', '-')}-population"
        resp = requests.get(url_wp, timeout=15)
        soup = BeautifulSoup(resp.text, "html.parser")
        table = soup.find("table")
        if table:
            rows = table.find_all("tr")[1:]
            city_pop_list = []
            for row in rows:
                cols = row.find_all("td")
                if len(cols) < 2:
                    continue
                city = cols[0].get_text(strip=True)
                pop_text = cols[1].get_text(strip=True).replace(",", "")
                try:
                    pop = int(pop_text)
                except:
                    continue
                city_pop_list.append((city, pop))
                if len(city_pop_list) >= MAX_CITIES:
                    break
            if city_pop_list:
                return city_pop_list
    except:
        pass

    # ---------------- CityPopulation.de ----------------
    try:
        url_cp = f"https://www.citypopulation.de/en/{country_name.replace(' ', '_')}/"
        resp = requests.get(url_cp, timeout=15)
        soup = BeautifulSoup(resp.text, "html.parser")
        table = soup.find("table", {"class": "data"})
        if table:
            rows = table.find_all("tr")[1:]
            city_pop_list = []
            for row in rows:
                cols = row.find_all("td")
                if len(cols) < 2:
                    continue
                city = cols[0].get_text(strip=True)
                pop_text = cols[1].get_text(strip=True).replace(",", "")
                try:
                    pop = int(pop_text)
                except:
                    continue
                city_pop_list.append((city, pop))
                if len(city_pop_list) >= MAX_CITIES:
                    break
            if city_pop_list:
                return city_pop_list
    except:
        pass

    # 都失败则返回未找到数据
    return [("未找到数据", 0)]

# ---------- 写入 Excel 并合并单元格 ----------
def write_excel(data, file_path):
    df = pd.DataFrame(data, columns=["城市", "人口", "国家", "语言", "时区", "洲"])
    df.to_excel(file_path, index=False)

    wb = load_workbook(file_path)
    ws = wb.active
    current_row = 2
    while current_row <= ws.max_row:
        country = ws.cell(current_row, 3).value
        start_row = current_row
        while current_row <= ws.max_row and ws.cell(current_row, 3).value == country:
            current_row += 1
        end_row = current_row - 1
        if end_row > start_row:
            for col in range(3, 7):  # 国家、语言、时区、洲列
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                ws.cell(start_row, col).alignment = Alignment(vertical="center", horizontal="center")
    wb.save(file_path)

# ---------- API ----------
@app.get("/generate_excel")
def generate_excel():
    rows = []
    countries = read_countries()
    if not countries:
        logging.warning("⚠️ countries.txt为空或未找到文件")
        return {"error": "countries.txt为空或未找到文件"}

    for cn_name in countries:
        cities = fetch_country_cities(cn_name)
        lang, tz, continent = COUNTRY_INFO.get(cn_name, ("未知", "未知", "未知"))
        for city, pop in cities:
            rows.append([city, pop, cn_name, lang, tz, continent])
        time.sleep(1)

    file_path = os.path.join(DATA_DIR, "country_population.xlsx")
    write_excel(rows, file_path)

    logging.info(f"✅ Excel 文件生成成功，可下载: {file_path}")

    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="country_population.xlsx"
    )

# ---------- 根路径 ----------
@app.get("/")
def root():
    return {"message": "API 正常运行，请访问 /generate_excel 下载 Excel 文件"}
